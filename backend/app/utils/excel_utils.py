"""
Excel Upload and Validation Utilities
"""

import pandas as pd
import io
from typing import Dict, List, Any, Tuple
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from fastapi import UploadFile, HTTPException
import re


class ExcelValidator:
    """Base class for Excel validation"""
    
    def __init__(self, schema: Dict[str, Dict[str, Any]]):
        """
        Initialize validator with schema
        
        Args:
            schema: Dictionary defining expected columns and their properties
                Example: {
                    'column_name': {
                        'required': True,
                        'type': str,
                        'validator': lambda x: x is not None,
                        'error_message': 'Custom error message'
                    }
                }
        """
        self.schema = schema
    
    def validate_file(self, file: UploadFile) -> Tuple[bool, str]:
        """Validate file extension"""
        if not file.filename.endswith(('.xlsx', '.xls')):
            return False, "Invalid file format. Please upload an Excel file (.xlsx or .xls)"
        return True, ""
    
    def validate_columns(self, df: pd.DataFrame) -> Tuple[bool, List[str]]:
        """Validate that all required columns are present"""
        errors = []
        required_columns = [col for col, props in self.schema.items() if props.get('required', True)]
        missing_columns = [col for col in required_columns if col not in df.columns]
        
        if missing_columns:
            errors.append(f"Missing required columns: {', '.join(missing_columns)}")
        
        return len(errors) == 0, errors
    
    def validate_data(self, df: pd.DataFrame) -> Tuple[bool, List[str]]:
        """Validate data types and values"""
        errors = []
        
        for idx, row in df.iterrows():
            row_num = idx + 2  # +2 because Excel rows start at 1 and we have a header
            
            for col, props in self.schema.items():
                if col not in df.columns:
                    continue
                
                value = row[col]
                
                # Check required fields
                if props.get('required', False) and pd.isna(value):
                    errors.append(f"Row {row_num}: '{col}' is required but empty")
                    continue
                
                # Skip validation if value is empty and not required
                if pd.isna(value) and not props.get('required', False):
                    continue
                
                # Type validation
                expected_type = props.get('type')
                if expected_type == 'email':
                    if not self._validate_email(str(value)):
                        errors.append(f"Row {row_num}: Invalid email format in '{col}'")
                elif expected_type == 'int':
                    try:
                        int(value)
                    except (ValueError, TypeError):
                        errors.append(f"Row {row_num}: '{col}' must be an integer")
                elif expected_type == 'float':
                    try:
                        float(value)
                    except (ValueError, TypeError):
                        errors.append(f"Row {row_num}: '{col}' must be a number")
                
                # Custom validator
                validator = props.get('validator')
                if validator and not validator(value):
                    error_msg = props.get('error_message', f"Invalid value in '{col}'")
                    errors.append(f"Row {row_num}: {error_msg}")
                
                # Min/Max validation for numbers
                if expected_type in ['int', 'float'] and not pd.isna(value):
                    try:
                        num_value = float(value)
                        if 'min' in props and num_value < props['min']:
                            errors.append(f"Row {row_num}: '{col}' must be at least {props['min']}")
                        if 'max' in props and num_value > props['max']:
                            errors.append(f"Row {row_num}: '{col}' must be at most {props['max']}")
                    except (ValueError, TypeError):
                        pass
                
                # Allowed values validation
                allowed_values = props.get('allowed_values')
                if allowed_values and value not in allowed_values:
                    errors.append(
                        f"Row {row_num}: '{col}' must be one of: {', '.join(map(str, allowed_values))}"
                    )
        
        return len(errors) == 0, errors
    
    def _validate_email(self, email: str) -> bool:
        """Validate email format"""
        pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
        return re.match(pattern, email) is not None
    
    async def validate_and_parse(self, file: UploadFile) -> Tuple[bool, List[Dict], List[str]]:
        """
        Main validation method
        
        Returns:
            Tuple of (is_valid, data_list, error_messages)
        """
        # Validate file extension
        is_valid, error = self.validate_file(file)
        if not is_valid:
            return False, [], [error]
        
        try:
            # Read Excel file
            contents = await file.read()
            df = pd.read_excel(io.BytesIO(contents))
            
            # Validate columns
            is_valid, errors = self.validate_columns(df)
            if not is_valid:
                return False, [], errors
            
            # Validate data
            is_valid, errors = self.validate_data(df)
            if not is_valid:
                return False, [], errors
            
            # Convert to list of dictionaries
            # Replace NaN with None
            df = df.where(pd.notna(df), None)
            
            # Special handling for integer columns - convert NaN to None instead of keeping float NaN
            for col in df.columns:
                if col in self.schema:
                    col_type = self.schema[col].get('type')
                    # For integer columns, explicitly convert NaN to None
                    if col_type == 'int':
                        df[col] = df[col].apply(lambda x: int(x) if pd.notna(x) and x is not None else None)
            
            data_list = df.to_dict('records')
            
            return True, data_list, []
            
        except Exception as e:
            return False, [], [f"Error reading Excel file: {str(e)}"]


class ExcelTemplateGenerator:
    """Generate Excel templates for data upload"""
    
    @staticmethod
    def generate_template(schema: Dict[str, Dict[str, Any]], template_name: str) -> bytes:
        """
        Generate an Excel template based on schema
        
        Args:
            schema: Dictionary defining columns and their properties
            template_name: Name of the template (will be shown in sheet name)
        
        Returns:
            bytes: Excel file content
        """
        wb = Workbook()
        ws = wb.active
        ws.title = template_name
        
        # Header styling
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Write headers
        headers = list(schema.keys())
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Write example row with placeholders
        for col_idx, (header, props) in enumerate(schema.items(), start=1):
            example_value = props.get('example', '')
            ws.cell(row=2, column=col_idx, value=example_value)
        
        # Add instructions sheet
        ws_instructions = wb.create_sheet(title="Instructions")
        ws_instructions.column_dimensions['A'].width = 30
        ws_instructions.column_dimensions['B'].width = 50
        
        instruction_title = ws_instructions.cell(row=1, column=1, value="Upload Instructions")
        instruction_title.font = Font(bold=True, size=14)
        
        ws_instructions.cell(row=3, column=1, value="Column Name")
        ws_instructions.cell(row=3, column=2, value="Description")
        ws_instructions.cell(row=3, column=3, value="Required")
        
        # Make instruction header bold
        for col in range(1, 4):
            cell = ws_instructions.cell(row=3, column=col)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        
        # Add column descriptions
        row_idx = 4
        for col_name, props in schema.items():
            ws_instructions.cell(row=row_idx, column=1, value=col_name)
            ws_instructions.cell(row=row_idx, column=2, value=props.get('description', ''))
            ws_instructions.cell(row=row_idx, column=3, value="Yes" if props.get('required', False) else "No")
            row_idx += 1
        
        # Adjust column widths
        for col_idx, header in enumerate(headers, start=1):
            ws.column_dimensions[chr(64 + col_idx)].width = 20
        
        # Save to bytes
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        return excel_buffer.getvalue()


# Schema definitions for different entities

CONTACT_SCHEMA = {
    'name': {
        'required': True,
        'type': 'str',
        'description': 'Full name of the contact',
        'example': 'John Doe'
    },
    'email': {
        'required': True,
        'type': 'email',
        'description': 'Email address',
        'example': 'john.doe@example.com'
    },
    'phone': {
        'required': False,
        'type': 'str',
        'description': 'Phone number',
        'example': '+1-234-567-8900'
    },
    'company': {
        'required': False,
        'type': 'str',
        'description': 'Company name',
        'example': 'Acme Corp'
    },
    'position': {
        'required': False,
        'type': 'str',
        'description': 'Job position/title',
        'example': 'Sales Manager'
    },
    'location': {
        'required': False,
        'type': 'str',
        'description': 'Location/Address',
        'example': 'New York, NY'
    },
    'status': {
        'required': False,
        'type': 'str',
        'allowed_values': ['Active', 'Inactive'],
        'description': 'Contact status',
        'example': 'Active'
    },
    'last_contact': {
        'required': False,
        'type': 'str',
        'description': 'Last contact date (YYYY-MM-DD)',
        'example': '2025-01-15'
    }
}

LEAD_SCHEMA = {
    'name': {
        'required': True,
        'type': 'str',
        'description': 'Lead name',
        'example': 'Jane Smith'
    },
    'company': {
        'required': True,
        'type': 'str',
        'description': 'Company name',
        'example': 'Tech Solutions Inc'
    },
    'email': {
        'required': True,
        'type': 'email',
        'description': 'Email address',
        'example': 'jane.smith@techsolutions.com'
    },
    'phone': {
        'required': False,
        'type': 'str',
        'description': 'Phone number',
        'example': '+1-234-567-8901'
    },
    'source': {
        'required': False,
        'type': 'str',
        'description': 'Lead source',
        'example': 'Website'
    },
    'status': {
        'required': False,
        'type': 'str',
        'allowed_values': ['New', 'Contacted', 'Qualified', 'Nurturing', 'Lost'],
        'description': 'Lead status',
        'example': 'New'
    },
    'score': {
        'required': False,
        'type': 'int',
        'min': 0,
        'max': 100,
        'description': 'Lead score (0-100)',
        'example': '75'
    },
    'value': {
        'required': False,
        'type': 'str',
        'description': 'Expected deal value',
        'example': '$50000'
    },
    'assigned_to': {
        'required': False,
        'type': 'str',
        'description': 'Assigned sales rep',
        'example': 'Sarah Johnson'
    }
}

OPPORTUNITY_SCHEMA = {
    'name': {
        'required': True,
        'type': 'str',
        'description': 'Opportunity name',
        'example': 'Enterprise Software Deal'
    },
    'account': {
        'required': True,
        'type': 'str',
        'description': 'Account/Company name',
        'example': 'Global Tech Corp'
    },
    'value': {
        'required': True,
        'type': 'float',
        'min': 0,
        'description': 'Deal value (numeric only)',
        'example': '150000'
    },
    'stage': {
        'required': False,
        'type': 'str',
        'allowed_values': ['Prospecting', 'Qualification', 'Proposal', 'Negotiation', 'Closed Won', 'Closed Lost'],
        'description': 'Sales stage',
        'example': 'Proposal'
    },
    'probability': {
        'required': False,
        'type': 'int',
        'min': 0,
        'max': 100,
        'description': 'Win probability (0-100)',
        'example': '60'
    },
    'close_date': {
        'required': False,
        'type': 'str',
        'description': 'Expected close date (YYYY-MM-DD)',
        'example': '2025-03-31'
    },
    'owner': {
        'required': False,
        'type': 'str',
        'description': 'Opportunity owner',
        'example': 'Mike Wilson'
    },
    'contact_id': {
        'required': False,
        'type': 'int',
        'description': 'Associated contact ID (optional)',
        'example': '123'
    }
}

ACCOUNT_SCHEMA = {
    'name': {
        'required': True,
        'type': 'str',
        'description': 'Account name',
        'example': 'Global Enterprises LLC'
    },
    'industry': {
        'required': False,
        'type': 'str',
        'description': 'Industry sector',
        'example': 'Technology'
    },
    'revenue': {
        'required': False,
        'type': 'str',
        'description': 'Annual revenue',
        'example': '$10M-50M'
    },
    'employees': {
        'required': False,
        'type': 'int',
        'min': 0,
        'description': 'Number of employees',
        'example': '500'
    },
    'location': {
        'required': False,
        'type': 'str',
        'description': 'Location/Address',
        'example': 'San Francisco, CA'
    },
    'phone': {
        'required': False,
        'type': 'str',
        'description': 'Main phone number',
        'example': '+1-415-555-0100'
    },
    'website': {
        'required': False,
        'type': 'str',
        'description': 'Company website',
        'example': 'https://www.example.com'
    },
    'account_owner': {
        'required': False,
        'type': 'str',
        'description': 'Account owner/manager',
        'example': 'Robert Brown'
    },
    'status': {
        'required': False,
        'type': 'str',
        'allowed_values': ['Active', 'Inactive', 'Prospect'],
        'description': 'Account status',
        'example': 'Active'
    }
}

TASK_SCHEMA = {
    'title': {
        'required': True,
        'type': 'str',
        'description': 'Task title',
        'example': 'Follow up with client'
    },
    'description': {
        'required': False,
        'type': 'str',
        'description': 'Task description',
        'example': 'Discuss Q1 renewal terms'
    },
    'priority': {
        'required': False,
        'type': 'str',
        'allowed_values': ['Low', 'Medium', 'High', 'Urgent'],
        'description': 'Task priority',
        'example': 'High'
    },
    'status': {
        'required': False,
        'type': 'str',
        'allowed_values': ['To Do', 'In Progress', 'Completed', 'Cancelled'],
        'description': 'Task status',
        'example': 'To Do'
    },
    'due_date': {
        'required': False,
        'type': 'str',
        'description': 'Due date (YYYY-MM-DD)',
        'example': '2025-01-20'
    },
    'assigned_to': {
        'required': False,
        'type': 'str',
        'description': 'Assigned to',
        'example': 'Alice Cooper'
    },
    'related_to': {
        'required': False,
        'type': 'str',
        'description': 'Related entity',
        'example': 'Acme Corp Deal'
    },
    'contact_id': {
        'required': False,
        'type': 'int',
        'description': 'Associated contact ID (optional)',
        'example': '123'
    }
}

EMAIL_CAMPAIGN_SCHEMA = {
    'name': {
        'required': True,
        'type': 'str',
        'description': 'Campaign name',
        'example': 'Q1 Product Launch'
    },
    'subject': {
        'required': True,
        'type': 'str',
        'description': 'Email subject line',
        'example': 'Introducing Our Latest Innovation'
    },
    'status': {
        'required': False,
        'type': 'str',
        'allowed_values': ['Draft', 'Scheduled', 'Sent', 'Paused'],
        'description': 'Campaign status',
        'example': 'Draft'
    },
    'sent_count': {
        'required': False,
        'type': 'int',
        'min': 0,
        'description': 'Number of emails sent',
        'example': '0'
    },
    'open_rate': {
        'required': False,
        'type': 'float',
        'min': 0,
        'max': 100,
        'description': 'Open rate percentage (0-100)',
        'example': '25.5'
    },
    'click_rate': {
        'required': False,
        'type': 'float',
        'min': 0,
        'max': 100,
        'description': 'Click rate percentage (0-100)',
        'example': '5.2'
    },
    'conversion_rate': {
        'required': False,
        'type': 'float',
        'min': 0,
        'max': 100,
        'description': 'Conversion rate percentage (0-100)',
        'example': '2.1'
    },
    'scheduled_date': {
        'required': False,
        'type': 'str',
        'description': 'Scheduled date (YYYY-MM-DD)',
        'example': '2025-02-01'
    }
}

CALENDAR_EVENT_SCHEMA = {
    'title': {
        'required': True,
        'type': 'str',
        'description': 'Event title',
        'example': 'Client Meeting'
    },
    'description': {
        'required': False,
        'type': 'str',
        'description': 'Event description',
        'example': 'Quarterly business review'
    },
    'event_type': {
        'required': False,
        'type': 'str',
        'allowed_values': ['Meeting', 'Call', 'Demo', 'Conference', 'Other'],
        'description': 'Type of event',
        'example': 'Meeting'
    },
    'start_time': {
        'required': True,
        'type': 'str',
        'description': 'Start time (YYYY-MM-DD HH:MM)',
        'example': '2025-01-20 14:00'
    },
    'end_time': {
        'required': False,
        'type': 'str',
        'description': 'End time (YYYY-MM-DD HH:MM)',
        'example': '2025-01-20 15:00'
    },
    'location': {
        'required': False,
        'type': 'str',
        'description': 'Location or meeting link',
        'example': 'Conference Room A'
    },
    'attendees': {
        'required': False,
        'type': 'str',
        'description': 'Attendees (comma-separated)',
        'example': 'john@example.com, jane@example.com'
    },
    'status': {
        'required': False,
        'type': 'str',
        'allowed_values': ['Scheduled', 'Completed', 'Cancelled', 'Rescheduled'],
        'description': 'Event status',
        'example': 'Scheduled'
    }
}
